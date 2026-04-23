"""
Step 1: Consolidate all sheets from edited_Sales_Funnel.xlsx into one flat CSV.

Output: lead_scoring_flat.csv
"""

import openpyxl
import csv
import re
from datetime import datetime

XLSX_PATH = "edited_Sales_Funnel.xlsx"
OUTPUT_PATH = "lead_scoring_flat.csv"

# ── Helpers ──────────────────────────────────────────────────────────────────

def clean(v):
    """Strip whitespace / newlines from a cell value."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace("\n", " ").replace("  ", " ")
        return v if v and v.lower() not in ("na", "n/a", "-", "--", "none") else None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v

def parse_stage_score(raw):
    """Extract numeric prefix from a sale stage string like '6.1 Installed LEAD.bot...'"""
    if raw is None:
        return None
    m = re.match(r"^(-?\d+(?:\.\d+)?)", str(raw).strip())
    return float(m.group(1)) if m else None

def stage_to_label(score):
    """Return (converted, outcome_class) from a numeric stage score."""
    if score is None:
        return 0, "unknown"
    if score >= 92:
        return 1, "paying"
    if score == -1:
        return 0, "canceled"
    if score in (0, 5):
        return 0, "inactive"
    # 1–91 = in-progress / warm leads; treat as not-yet-converted
    return 0, "in_progress"

FIELDNAMES = [
    "sheet_origin",
    "job_title",
    "company",
    "industry",
    "platform",
    "lead_id",
    "company_size_numeric",
    "company_size_range",
    "region",
    "trial",
    "sale_stage_raw",
    "sale_stage_score",
    "first_contact_date",
    "last_contact_date",
    "current_plan",
    "payment_value_2024",
    "payment_value_2023",
    "nda",
    "plan_cancel_date",
    "active_users",
    "engagement_rate",
    "converted",
    "outcome_class",
]

rows_out = []

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# ── Sheet 1: Inbound Sales leads ─────────────────────────────────────────────
# col: 0=job_title 1=phone 2=company 3=industry 4=platform 5=id
#      6=first_contact 7=last_contact 8=size_num 9=size_range 10=trial
#      11=sale_stage 12=current_status 13=comment 14=sales_note
#      15=region 16=nda 17=unsub 18=reason 19=fd_notes

ws = wb[" Inbound Sales leads"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not any(r[i] for i in (0, 2, 3, 11)):
        continue
    stage_raw = clean(r[11])
    score = parse_stage_score(stage_raw)
    converted, outcome = stage_to_label(score)
    rows_out.append({
        "sheet_origin": "inbound",
        "job_title": clean(r[0]),
        "company": clean(r[2]),
        "industry": clean(r[3]),
        "platform": clean(r[4]),
        "lead_id": clean(r[5]),
        "company_size_numeric": clean(r[8]),
        "company_size_range": clean(r[9]),
        "region": clean(r[15]),
        "trial": 1 if clean(r[10]) else 0,
        "sale_stage_raw": stage_raw,
        "sale_stage_score": score,
        "first_contact_date": clean(r[6]),
        "last_contact_date": clean(r[7]),
        "current_plan": None,
        "payment_value_2024": None,
        "payment_value_2023": None,
        "nda": 1 if str(clean(r[16]) or "").lower() in ("yes", "true", "1") else 0,
        "plan_cancel_date": None,
        "active_users": None,
        "engagement_rate": None,
        "converted": converted,
        "outcome_class": outcome,
    })

# ── Sheet 2: LEAD pay user ────────────────────────────────────────────────────
# col: 4=job_title 6=company 7=industry 8=size_num 9=size_range 10=platform
#      11=id 12=current_plan 13=sale_stage 14=last_contact 15=payment_2024
#      16=payment_method 17=next_renewal 18=service_start 19=latest_payment
#      20=monthly_$ 21=expand_opp 22=sales_note 23=customer_since
#      26=nda 28=payment_2023 30=active_users 31=engagement_rate 35=region

ws = wb["LEAD pay user"]
for r in ws.iter_rows(min_row=3, values_only=True):
    if not any(r[i] for i in (4, 6, 13)):
        continue
    stage_raw = clean(r[13])
    score = parse_stage_score(stage_raw)
    converted, outcome = stage_to_label(score)
    # Pay users default to converted=1 if stage not parseable
    if score is None or score >= 90:
        converted, outcome = 1, "paying"
    rows_out.append({
        "sheet_origin": "pay_user",
        "job_title": clean(r[4]),
        "company": clean(r[6]),
        "industry": clean(r[7]),
        "platform": clean(r[10]),
        "lead_id": clean(r[11]),
        "company_size_numeric": clean(r[8]),
        "company_size_range": clean(r[9]),
        "region": clean(r[35]),
        "trial": 0,
        "sale_stage_raw": stage_raw,
        "sale_stage_score": score,
        "first_contact_date": clean(r[23]),   # customer_since as proxy
        "last_contact_date": clean(r[14]),
        "current_plan": clean(r[12]),
        "payment_value_2024": clean(r[15]),
        "payment_value_2023": clean(r[28]),
        "nda": 1 if str(clean(r[26]) or "").lower() in ("yes", "true", "1") else 0,
        "plan_cancel_date": None,
        "active_users": clean(r[30]),
        "engagement_rate": clean(r[31]),
        "converted": converted,
        "outcome_class": outcome,
    })

# ── Sheet 3: Canceled plan ────────────────────────────────────────────────────
# col: 3=job_title 5=company 6=industry 7=size_range 8=plan_size_2022
#      9=platform 10=sale_stage 11=last_contact 12=status 13=cancel_date
#      16=sales_note

ws = wb["Canceled plan"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not any(r[i] for i in (3, 5, 10)):
        continue
    stage_raw = clean(r[10])
    score = parse_stage_score(stage_raw)
    rows_out.append({
        "sheet_origin": "canceled",
        "job_title": clean(r[3]),
        "company": clean(r[5]),
        "industry": clean(r[6]),
        "platform": clean(r[9]),
        "lead_id": None,
        "company_size_numeric": None,
        "company_size_range": clean(r[7]),
        "region": None,
        "trial": 0,
        "sale_stage_raw": stage_raw,
        "sale_stage_score": score if score is not None else -1,
        "first_contact_date": None,
        "last_contact_date": clean(r[11]),
        "current_plan": clean(r[8]),  # plan size 2022
        "payment_value_2024": None,
        "payment_value_2023": None,
        "nda": 0,
        "plan_cancel_date": clean(r[13]),
        "active_users": None,
        "engagement_rate": None,
        "converted": 0,
        "outcome_class": "canceled",
    })

# ── Sheet 4: Inactive Lead ────────────────────────────────────────────────────
# col: 3=job_title 4=additional 5=company 6=industry 7=platform 8=id
#      9=size 10=first_contact 11=last_contact 12=sale_stage
#      13=current_note 14=comments 20=unsubscribe 21=reason

ws = wb["Inactive Lead"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not any(r[i] for i in (3, 5, 12)):
        continue
    # skip rows where col 3 looks like an email (header confusion rows)
    jt = clean(r[3])
    if jt and "@" in jt:
        jt = None
    stage_raw = clean(r[12])
    score = parse_stage_score(stage_raw)
    rows_out.append({
        "sheet_origin": "inactive",
        "job_title": jt,
        "company": clean(r[5]),
        "industry": clean(r[6]),
        "platform": clean(r[7]),
        "lead_id": clean(r[8]),
        "company_size_numeric": clean(r[9]),
        "company_size_range": None,
        "region": None,
        "trial": 0,
        "sale_stage_raw": stage_raw,
        "sale_stage_score": score if score is not None else 0,
        "first_contact_date": clean(r[10]),
        "last_contact_date": clean(r[11]),
        "current_plan": None,
        "payment_value_2024": None,
        "payment_value_2023": None,
        "nda": 0,
        "plan_cancel_date": None,
        "active_users": None,
        "engagement_rate": None,
        "converted": 0,
        "outcome_class": "inactive",
    })

# ── Write output ──────────────────────────────────────────────────────────────

with open(OUTPUT_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
    writer.writeheader()
    writer.writerows(rows_out)

# ── Summary ───────────────────────────────────────────────────────────────────

from collections import Counter
origins = Counter(r["sheet_origin"] for r in rows_out)
outcomes = Counter(r["outcome_class"] for r in rows_out)
converted = sum(r["converted"] for r in rows_out)

print(f"Total rows written: {len(rows_out)}")
print(f"By sheet:   {dict(origins)}")
print(f"By outcome: {dict(outcomes)}")
print(f"Converted=1: {converted} ({100*converted/len(rows_out):.1f}%)")
print(f"Output: {OUTPUT_PATH}")
